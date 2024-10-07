import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

bg_colors = {
    'light_blue': PatternFill(start_color='DDE9F6', end_color='DDE9F6', fill_type='solid'),
    'dark_gray': PatternFill(start_color='393939', end_color='393939', fill_type='solid'),
    'medium_dark_gray': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
    'light_gray': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
}
lb_headers = ['inflammation', 'large_interstitial', 'papillary_dermis', 'reticular_dermis']
dg_headers = ['no_label']; mdg_headers = ['endothelial', 'vessel_wall']; lg_headers = ['unknown', 'shade', 'noise']

def fill_rows_cols(ws, percentage_df, column_headers, fill_color):
    for header in column_headers:
        if header in percentage_df.columns:
            col_index = percentage_df.columns.get_loc(header) + 2
            ws.cell(row=1, column=col_index).fill = fill_color
                
        if header in percentage_df.index:
            row_index = percentage_df.index.get_loc(header) + 2
            for col_num in range(1, ws.max_column + 1):  
                ws.cell(row=row_index, column=col_num).fill = fill_color
        
        if header in percentage_df.index:
            row_index = percentage_df.index.get_loc(header) + 2
               
def red_text(ws, percentage_df):
    red_headers = ['RBC', 'background', 'ecm']
    for header in red_headers:
        if header in percentage_df.columns:
            col_index = percentage_df.columns.get_loc(header) + 2  
            ws.cell(row=1, column=col_index).font = Font(color="FF0000")  
            
        if header in percentage_df.index:
            row_index = percentage_df.index.get_loc(header) + 2 
            ws.cell(row=row_index, column=1).font = Font(color="FF0000") 

    for i in range(2, ws.max_row + 1): 
        if i <= ws.max_column:
            ws.cell(row=i, column=i).font = Font(color="FF0000", bold=True)  
        
###############################################################################################

def percent_matrix(merged_matrix_file, isTest):
    df = pd.read_excel(merged_matrix_file, sheet_name=0, index_col=0, header=0)

    row_sums = df.sum(axis = 1)

    percentage_df = df.div(row_sums, axis=0) 

    # if(isTest == True):
    #     print("Percentage Matrix TEST:")
    # else:
    #     print("Percentage Matrix TRAIN:")
    # print(percentage_df)

    # Excel Formatting
    output_title = 'Percentage_matrix_train'  
    if(isTest == True):
        output_title = 'Percentage_matrix_test' 
    output_file_path = f'./Outputs/{output_title}.xlsx'
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        percentage_df.to_excel(writer, sheet_name=f'{output_title}', index=True, header=True)
    wb = load_workbook(output_file_path)
    ws = wb[f'{output_title}']

    # fill bg
    fill_rows_cols(ws, percentage_df, mdg_headers, bg_colors['medium_dark_gray'])
    fill_rows_cols(ws, percentage_df, lg_headers, bg_colors['light_gray'])
    fill_rows_cols(ws, percentage_df, lb_headers, bg_colors['light_blue'])
    fill_rows_cols(ws, percentage_df, dg_headers, bg_colors['dark_gray'])

    # color the texts to red
    red_text(ws, percentage_df)
    
    # save formatting changes
    wb.save(output_file_path)
    
    return output_file_path # Return the file path