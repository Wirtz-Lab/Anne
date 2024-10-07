# TASK 2
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

#bg colors
bg_colors = {
    'light_blue': PatternFill(start_color='DDE9F6', end_color='DDE9F6', fill_type='solid'),
    'dark_gray': PatternFill(start_color='393939', end_color='393939', fill_type='solid'),
    'medium_dark_gray': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
    'light_gray': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
}
lb_headers = ['inflammation', 'large_interstitial', 'papillary_dermis', 'reticular_dermis']
dg_headers = ['no_label']; mdg_headers = ['endothelial', 'vessel_wall']; lg_headers = ['unknown', 'shade', 'noise']


def fill_rows_cols(ws, merged_df, column_headers, fill_color):
    for header in column_headers:
        if header in merged_df.columns:
            col_index = merged_df.columns.get_loc(header) + 2
            for row_num in range(1, ws.max_row + 1): 
                ws.cell(row=row_num, column=col_index).fill = fill_color
                
        if header in merged_df.index:
            row_index = merged_df.index.get_loc(header) + 2
            for col_num in range(1, ws.max_column + 1):  
                ws.cell(row=row_index, column=col_num).fill = fill_color
                
def red_text(ws, merged_df, red_headers):
    for header in red_headers:
        if header in merged_df.columns:
            col_index = merged_df.columns.get_loc(header) + 2  
            ws.cell(row=1, column=col_index).font = Font(color="FF0000")  
            
        if header in merged_df.index:
            row_index = merged_df.index.get_loc(header) + 2  
            ws.cell(row=row_index, column=1).font = Font(color="FF0000")  

###################################################################

def merge_matrix(sum_matrix_file, isTest):
    df = pd.read_excel(sum_matrix_file, sheet_name=0, index_col=0, header=0)
    merged_df = df.copy() # cp the dataframe

    merged_df['RBC'] = merged_df['RBC'] + merged_df['endothelial'] + merged_df['vessel_wall'] 
    merged_df['endothelial'] = np.nan; merged_df.loc['endothelial',:] = np.nan
    merged_df['vessel_wall'] = np.nan; merged_df.loc['vessel_wall',:] = np.nan

    merged_df['background'] = merged_df['background'] + merged_df['shade'] + merged_df['noise'] + merged_df['unknown'] 
    merged_df['shade'] = np.nan; merged_df.loc['shade',:] = np.nan
    merged_df['noise'] = np.nan; merged_df.loc['noise',:] = np.nan
    merged_df['unknown'] = np.nan; merged_df.loc['unknown',:] = np.nan

    merged_df['ecm'] = merged_df['ecm'] + merged_df['inflammation'] + merged_df['large_interstitial'] + merged_df['papillary_dermis'] + merged_df['reticular_dermis']
    merged_df['inflammation'] = np.nan; merged_df.loc['inflammation',:] = np.nan
    merged_df['large_interstitial'] = np.nan; merged_df.loc['large_interstitial',:] = np.nan
    merged_df['papillary_dermis'] = np.nan; merged_df.loc['papillary_dermis',:] = np.nan
    merged_df['reticular_dermis'] = np.nan; merged_df.loc['reticular_dermis',:] = np.nan

    merged_df['no_label'] = np.nan; merged_df.loc['no_label',:] = np.nan
    
    # if(isTest == True):
    #     print("Merged Matrix TEST:")
    # else:
    #     print("Merged Matrix TRAIN:")
    # print(merged_df)

    # Excel Formatting
    output_title = 'Merged_matrix_train'  
    if(isTest == True):
        output_title = 'Merged_matrix_test' 
    output_file_path = f'./Outputs/{output_title}.xlsx'
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name=f'{output_title}', index=True, header=True)
    wb = load_workbook(output_file_path)
    ws = wb[f'{output_title}']

    # fill bg
    fill_rows_cols(ws, merged_df, mdg_headers, bg_colors['medium_dark_gray'])
    fill_rows_cols(ws, merged_df, lg_headers, bg_colors['light_gray'])
    fill_rows_cols(ws, merged_df, lb_headers, bg_colors['light_blue'])
    fill_rows_cols(ws, merged_df, dg_headers, bg_colors['dark_gray'])

    # red color for merged headers
    red_headers = ['RBC', 'background', 'ecm']
    red_text(ws, merged_df, red_headers)
        
    wb.save(output_file_path)
    
    return output_file_path # Return the file path
        

