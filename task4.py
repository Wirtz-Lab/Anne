# TASK 4
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

#bg colors
bg_colors = {
    'light_blue': PatternFill(start_color='DDE9F6', end_color='DDE9F6', fill_type='solid'),
    'dark_gray': PatternFill(start_color='393939', end_color='393939', fill_type='solid'),
    'medium_dark_gray': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
    'light_gray': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
}
lb_headers = ['inflammation', 'large_interstitial', 'papillary_dermis', 'reticular_dermis']
dg_headers = ['no_label']; mdg_headers = ['endothelial', 'vessel_wall']; lg_headers = ['unknown', 'shade', 'noise']


def fill_rows_cols(ws, precision_df, headers, fill_color):
    for header in headers:
        if header in precision_df.index:
            row_index = precision_df.index.get_loc(header) + 2 
            ws.cell(row=row_index, column=1).fill = fill_color  

###################################################################

def precision_matrix(percent_train_file, percent_test_file):
    df_train = pd.read_excel(percent_train_file, sheet_name=0, index_col=0, header=0)
    df_test = pd.read_excel(percent_test_file, sheet_name=0, index_col=0, header=0)

    precision_df = pd.DataFrame(0.0, columns = ['Train', 'Test'], index = df_test.index)

    print(df_train)

    # Extracting diagonal values
    for i in range(min(len(df_train), len(df_train.columns))):
        precision_df.at[df_train.index[i], 'Train'] = df_train.iat[i, i]
    for i in range(min(len(df_test), len(df_test.columns))):
        precision_df.at[df_test.index[i], 'Test'] = df_test.iat[i, i]
        
    # Adding average of all classes 
    avg_precision = precision_df.sum() / precision_df.index.size * 100
    avg_precision.name = 'average'
    precision_df = pd.concat([precision_df, avg_precision.to_frame().T])
    precision_df = precision_df.fillna(0)

    print(precision_df)

    # Excel Formatting
    output_file_path = './Precision.xlsx'  
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        precision_df.to_excel(writer, sheet_name='Precision', index=True, header=True)
    wb = load_workbook(output_file_path)
    ws = wb['Precision']

    # fill bg
    fill_rows_cols(ws, precision_df, mdg_headers, bg_colors['medium_dark_gray'])
    fill_rows_cols(ws, precision_df, lg_headers, bg_colors['light_gray'])
    fill_rows_cols(ws, precision_df, lb_headers, bg_colors['light_blue'])
    fill_rows_cols(ws, precision_df, dg_headers, bg_colors['dark_gray'])

    # adding bold font
    for col_num in range(1, ws.max_column + 1):  # Iterate over all columns
        ws.cell(row=precision_df.index.size + 1, column=col_num).font = Font(bold=True)

    wb.save(output_file_path)
    
    